import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, date

def create_transport_management_excel():
    """運送会社向け経営管理Excelファイルを作成"""
    
    # ワークブック作成
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除
    
    # カラーパレット定義
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=12)
    
    # 罫線スタイル
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # =========================
    # 1. マスタシート作成
    # =========================
    ws_master = wb.create_sheet("マスタ")
    ws_master.sheet_view.showGridLines = False
    
    # タイトル
    ws_master['A1'] = "マスタデータ管理"
    ws_master['A1'].font = title_font
    ws_master.merge_cells('A1:F1')
    
    # 取引先マスタ
    ws_master['A3'] = "取引先一覧"
    ws_master['A3'].font = subtitle_font
    ws_master['A4'] = "取引先コード"
    ws_master['B4'] = "取引先名"
    ws_master['A4'].fill = header_fill
    ws_master['A4'].font = header_font
    ws_master['B4'].fill = header_fill
    ws_master['B4'].font = header_font
    
    clients = [
        ["C001", "A運送株式会社"],
        ["C002", "B物流センター"],
        ["C003", "C製造工場"],
        ["C004", "D商事"],
        ["C005", "E倉庫"]
    ]
    for i, client in enumerate(clients, 5):
        ws_master[f'A{i}'] = client[0]
        ws_master[f'B{i}'] = client[1]
    
    # 車両マスタ
    ws_master['D3'] = "車両番号一覧"
    ws_master['D3'].font = subtitle_font
    ws_master['D4'] = "車両番号"
    ws_master['E4'] = "車種"
    ws_master['D4'].fill = header_fill
    ws_master['D4'].font = header_font
    ws_master['E4'].fill = header_fill
    ws_master['E4'].font = header_font
    
    vehicles = [
        ["車両001", "2tトラック"],
        ["車両002", "4tトラック"],
        ["車両003", "4tトラック"],
        ["車両004", "10tトラック"],
        ["車両005", "10tトラック"]
    ]
    for i, vehicle in enumerate(vehicles, 5):
        ws_master[f'D{i}'] = vehicle[0]
        ws_master[f'E{i}'] = vehicle[1]
    
    # ドライバーマスタ
    ws_master['G3'] = "ドライバー一覧"
    ws_master['G3'].font = subtitle_font
    ws_master['G4'] = "社員番号"
    ws_master['H4'] = "ドライバー名"
    ws_master['G4'].fill = header_fill
    ws_master['G4'].font = header_font
    ws_master['H4'].fill = header_fill
    ws_master['H4'].font = header_font
    
    drivers = [
        ["D001", "田中太郎"],
        ["D002", "佐藤次郎"],
        ["D003", "鈴木三郎"],
        ["D004", "高橋四郎"],
        ["D005", "山田五郎"]
    ]
    for i, driver in enumerate(drivers, 5):
        ws_master[f'G{i}'] = driver[0]
        ws_master[f'H{i}'] = driver[1]
    
    # 列幅調整
    ws_master.column_dimensions['A'].width = 15
    ws_master.column_dimensions['B'].width = 20
    ws_master.column_dimensions['D'].width = 15
    ws_master.column_dimensions['E'].width = 20
    ws_master.column_dimensions['G'].width = 15
    ws_master.column_dimensions['H'].width = 20
    
    # =========================
    # 2. 売上入力シート
    # =========================
    ws_sales = wb.create_sheet("売上入力")
    ws_sales.sheet_view.showGridLines = False
    
    ws_sales['A1'] = "売上データ入力"
    ws_sales['A1'].font = title_font
    ws_sales.merge_cells('A1:F1')
    
    # ヘッダー設定
    headers = ["日付", "月", "取引先名", "車両番号", "売上金額", "備考"]
    for col, header in enumerate(headers, 1):
        cell = ws_sales.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # データ検証（プルダウン）設定
    dv_client = DataValidation(type="list", formula1="=マスタ!$B$5:$B$9", allow_blank=True)
    dv_vehicle = DataValidation(type="list", formula1="=マスタ!$D$5:$D$9", allow_blank=True)
    
    ws_sales.add_data_validation(dv_client)
    ws_sales.add_data_validation(dv_vehicle)
    
    # サンプルデータと数式
    sample_sales = [
        [date(2025, 1, 5), "", "A運送株式会社", "車両001", 150000, "定期便"],
        [date(2025, 1, 10), "", "B物流センター", "車両002", 200000, "特別便"],
        [date(2025, 1, 15), "", "C製造工場", "車両003", 180000, ""],
    ]
    
    for row_idx, data in enumerate(sample_sales, 4):
        ws_sales.cell(row=row_idx, column=1, value=data[0]).number_format = 'yyyy/mm/dd'
        ws_sales.cell(row=row_idx, column=2, value=f'=MONTH(A{row_idx})')  # 月を自動抽出
        ws_sales.cell(row=row_idx, column=3, value=data[2])
        ws_sales.cell(row=row_idx, column=4, value=data[3])
        ws_sales.cell(row=row_idx, column=5, value=data[4]).number_format = '#,##0'
        ws_sales.cell(row=row_idx, column=6, value=data[5])
        
        # プルダウン適用
        dv_client.add(ws_sales.cell(row=row_idx, column=3))
        dv_vehicle.add(ws_sales.cell(row=row_idx, column=4))
    
    # 列幅調整
    ws_sales.column_dimensions['A'].width = 12
    ws_sales.column_dimensions['B'].width = 8
    ws_sales.column_dimensions['C'].width = 20
    ws_sales.column_dimensions['D'].width = 15
    ws_sales.column_dimensions['E'].width = 15
    ws_sales.column_dimensions['F'].width = 20
    
    # =========================
    # 3. 経費入力シート
    # =========================
    ws_expense = wb.create_sheet("経費入力")
    ws_expense.sheet_view.showGridLines = False
    
    ws_expense['A1'] = "経費データ入力"
    ws_expense['A1'].font = title_font
    ws_expense.merge_cells('A1:F1')
    
    # ヘッダー設定
    headers = ["日付", "月", "車両番号", "経費区分", "金額", "備考"]
    for col, header in enumerate(headers, 1):
        cell = ws_expense.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # 経費区分のプルダウン
    expense_types = "燃料費,修理費,保険料,リース料,その他"
    dv_expense = DataValidation(type="list", formula1=f'"{expense_types}"', allow_blank=True)
    ws_expense.add_data_validation(dv_expense)
    
    # サンプルデータ
    sample_expenses = [
        [date(2025, 1, 5), "", "車両001", "燃料費", 25000, "軽油"],
        [date(2025, 1, 10), "", "車両002", "修理費", 50000, "タイヤ交換"],
        [date(2025, 1, 15), "", "車両001", "保険料", 30000, ""],
    ]
    
    for row_idx, data in enumerate(sample_expenses, 4):
        ws_expense.cell(row=row_idx, column=1, value=data[0]).number_format = 'yyyy/mm/dd'
        ws_expense.cell(row=row_idx, column=2, value=f'=MONTH(A{row_idx})')
        ws_expense.cell(row=row_idx, column=3, value=data[2])
        ws_expense.cell(row=row_idx, column=4, value=data[3])
        ws_expense.cell(row=row_idx, column=5, value=data[4]).number_format = '#,##0'
        ws_expense.cell(row=row_idx, column=6, value=data[5])
        
        dv_vehicle.add(ws_expense.cell(row=row_idx, column=3))
        dv_expense.add(ws_expense.cell(row=row_idx, column=4))
    
    # 列幅調整
    ws_expense.column_dimensions['A'].width = 12
    ws_expense.column_dimensions['B'].width = 8
    ws_expense.column_dimensions['C'].width = 15
    ws_expense.column_dimensions['D'].width = 15
    ws_expense.column_dimensions['E'].width = 15
    ws_expense.column_dimensions['F'].width = 20
    
    # =========================
    # 4. 人件費入力シート
    # =========================
    ws_labor = wb.create_sheet("人件費入力")
    ws_labor.sheet_view.showGridLines = False
    
    ws_labor['A1'] = "人件費データ入力"
    ws_labor['A1'].font = title_font
    ws_labor.merge_cells('A1:G1')
    
    # ヘッダー設定
    headers = ["ドライバー名", "車両番号", "按分比率(%)", "月", "支給額", "按分額", "備考"]
    for col, header in enumerate(headers, 1):
        cell = ws_labor.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # データ検証
    dv_driver = DataValidation(type="list", formula1="=マスタ!$H$5:$H$9", allow_blank=True)
    ws_labor.add_data_validation(dv_driver)
    
    # サンプルデータ
    sample_labor = [
        ["田中太郎", "車両001", 100, 1, 300000, "", ""],
        ["佐藤次郎", "車両002", 100, 1, 280000, "", ""],
        ["鈴木三郎", "車両003", 50, 1, 260000, "", "複数車両担当"],
        ["鈴木三郎", "車両004", 50, 1, 260000, "", "複数車両担当"],
    ]
    
    for row_idx, data in enumerate(sample_labor, 4):
        ws_labor.cell(row=row_idx, column=1, value=data[0])
        ws_labor.cell(row=row_idx, column=2, value=data[1])
        ws_labor.cell(row=row_idx, column=3, value=data[2]/100).number_format = '0%'
        ws_labor.cell(row=row_idx, column=4, value=data[3])
        ws_labor.cell(row=row_idx, column=5, value=data[4]).number_format = '#,##0'
        ws_labor.cell(row=row_idx, column=6, value=f'=E{row_idx}*C{row_idx}').number_format = '#,##0'
        ws_labor.cell(row=row_idx, column=7, value=data[6])
        
        dv_driver.add(ws_labor.cell(row=row_idx, column=1))
        dv_vehicle.add(ws_labor.cell(row=row_idx, column=2))
    
    # 列幅調整
    ws_labor.column_dimensions['A'].width = 15
    ws_labor.column_dimensions['B'].width = 15
    ws_labor.column_dimensions['C'].width = 12
    ws_labor.column_dimensions['D'].width = 8
    ws_labor.column_dimensions['E'].width = 15
    ws_labor.column_dimensions['F'].width = 15
    ws_labor.column_dimensions['G'].width = 20
    
    # =========================
    # 5. 資金繰り入力シート
    # =========================
    ws_cash = wb.create_sheet("資金繰り入力")
    ws_cash.sheet_view.showGridLines = False
    
    ws_cash['A1'] = "資金繰りデータ入力"
    ws_cash['A1'].font = title_font
    ws_cash.merge_cells('A1:F1')
    
    # ヘッダー設定
    headers = ["日付", "月", "入金/出金", "科目", "金額", "備考"]
    for col, header in enumerate(headers, 1):
        cell = ws_cash.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # プルダウン設定
    dv_inout = DataValidation(type="list", formula1='"入金,出金"', allow_blank=True)
    dv_category = DataValidation(type="list", formula1='"売上入金,燃料費支払,リース料支払,借入返済,利息,その他"', allow_blank=True)
    ws_cash.add_data_validation(dv_inout)
    ws_cash.add_data_validation(dv_category)
    
    # サンプルデータ
    sample_cash = [
        [date(2025, 1, 1), "", "入金", "売上入金", 500000, "前月売上回収"],
        [date(2025, 1, 10), "", "出金", "燃料費支払", 150000, ""],
        [date(2025, 1, 20), "", "出金", "リース料支払", 200000, "車両リース"],
    ]
    
    for row_idx, data in enumerate(sample_cash, 4):
        ws_cash.cell(row=row_idx, column=1, value=data[0]).number_format = 'yyyy/mm/dd'
        ws_cash.cell(row=row_idx, column=2, value=f'=MONTH(A{row_idx})')
        ws_cash.cell(row=row_idx, column=3, value=data[2])
        ws_cash.cell(row=row_idx, column=4, value=data[3])
        ws_cash.cell(row=row_idx, column=5, value=data[4]).number_format = '#,##0'
        ws_cash.cell(row=row_idx, column=6, value=data[5])
        
        dv_inout.add(ws_cash.cell(row=row_idx, column=3))
        dv_category.add(ws_cash.cell(row=row_idx, column=4))
    
    # 列幅調整
    ws_cash.column_dimensions['A'].width = 12
    ws_cash.column_dimensions['B'].width = 8
    ws_cash.column_dimensions['C'].width = 12
    ws_cash.column_dimensions['D'].width = 18
    ws_cash.column_dimensions['E'].width = 15
    ws_cash.column_dimensions['F'].width = 20
    
    # =========================
    # 6. 損益計算表
    # =========================
    ws_pl = wb.create_sheet("損益計算表")
    ws_pl.sheet_view.showGridLines = False
    
    ws_pl['A1'] = "月次損益計算表"
    ws_pl['A1'].font = title_font
    ws_pl.merge_cells('A1:P1')
    
    # 月ヘッダー
    ws_pl['A3'] = "項目"
    ws_pl['A3'].fill = header_fill
    ws_pl['A3'].font = header_font
    
    months = ["1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "年間合計"]
    for col, month in enumerate(months, 2):
        cell = ws_pl.cell(row=3, column=col, value=month)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # 損益項目
    pl_items = [
        "売上高（予算）",
        "売上高（実績）",
        "差異",
        "",
        "燃料費",
        "修理費",
        "保険料",
        "リース料",
        "人件費",
        "その他経費",
        "経費合計",
        "",
        "営業利益",
        "営業利益率"
    ]
    
    for row_idx, item in enumerate(pl_items, 4):
        ws_pl.cell(row=row_idx, column=1, value=item).font = Font(bold=True if item in ["売上高（実績）", "経費合計", "営業利益"] else False)
    
    # 数式設定（1月の例）
    # 売上高実績
    ws_pl['B5'] = '=SUMIFS(売上入力!E:E,売上入力!B:B,1)'
    ws_pl['B5'].number_format = '#,##0'
    
    # 予算（前年実績の110%と仮定）
    ws_pl['B4'] = 500000
    ws_pl['B4'].number_format = '#,##0'
    
    # 差異
    ws_pl['B6'] = '=B5-B4'
    ws_pl['B6'].number_format = '#,##0'
    
    # 経費項目
    ws_pl['B8'] = '=SUMIFS(経費入力!E:E,経費入力!B:B,1,経費入力!D:D,"燃料費")'
    ws_pl['B9'] = '=SUMIFS(経費入力!E:E,経費入力!B:B,1,経費入力!D:D,"修理費")'
    ws_pl['B10'] = '=SUMIFS(経費入力!E:E,経費入力!B:B,1,経費入力!D:D,"保険料")'
    ws_pl['B11'] = '=SUMIFS(経費入力!E:E,経費入力!B:B,1,経費入力!D:D,"リース料")'
    ws_pl['B12'] = '=SUMIFS(人件費入力!F:F,人件費入力!D:D,1)'
    ws_pl['B13'] = '=SUMIFS(経費入力!E:E,経費入力!B:B,1,経費入力!D:D,"その他")'
    ws_pl['B14'] = '=SUM(B8:B13)'
    
    # 営業利益
    ws_pl['B16'] = '=B5-B14'
    ws_pl['B17'] = '=IF(B5=0,0,B16/B5)'
    ws_pl['B17'].number_format = '0.0%'
    
    # 条件付き書式（営業利益がマイナスの場合赤文字）
    red_font = Font(color="FF0000")
    ws_pl.conditional_formatting.add('B16:N16',
        CellIsRule(operator='lessThan', formula=['0'], font=red_font))
    
    # 列幅調整
    ws_pl.column_dimensions['A'].width = 20
    for col in range(2, 15):
        ws_pl.column_dimensions[chr(64 + col)].width = 12
    
    # =========================
    # 7. 車両別収支
    # =========================
    ws_vehicle = wb.create_sheet("車両別収支")
    ws_vehicle.sheet_view.showGridLines = False
    
    ws_vehicle['A1'] = "車両別収支一覧"
    ws_vehicle['A1'].font = title_font
    ws_vehicle.merge_cells('A1:F1')
    
    # ヘッダー
    headers = ["車両番号", "売上高", "経費", "人件費", "利益", "利益率"]
    for col, header in enumerate(headers, 1):
        cell = ws_vehicle.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # 車両データ
    vehicles = ["車両001", "車両002", "車両003", "車両004", "車両005"]
    for row_idx, vehicle in enumerate(vehicles, 4):
        ws_vehicle.cell(row=row_idx, column=1, value=vehicle)
        # 売上高
        ws_vehicle.cell(row=row_idx, column=2, value=f'=SUMIF(売上入力!D:D,A{row_idx},売上入力!E:E)').number_format = '#,##0'
        # 経費
        ws_vehicle.cell(row=row_idx, column=3, value=f'=SUMIF(経費入力!C:C,A{row_idx},経費入力!E:E)').number_format = '#,##0'
        # 人件費
        ws_vehicle.cell(row=row_idx, column=4, value=f'=SUMIF(人件費入力!B:B,A{row_idx},人件費入力!F:F)').number_format = '#,##0'
        # 利益
        ws_vehicle.cell(row=row_idx, column=5, value=f'=B{row_idx}-C{row_idx}-D{row_idx}').number_format = '#,##0'
        # 利益率
        ws_vehicle.cell(row=row_idx, column=6, value=f'=IF(B{row_idx}=0,0,E{row_idx}/B{row_idx})').number_format = '0.0%'
    
    # 合計行
    ws_vehicle.cell(row=10, column=1, value="合計").font = Font(bold=True)
    for col in range(2, 6):
        ws_vehicle.cell(row=10, column=col, value=f'=SUM({chr(64+col)}4:{chr(64+col)}8)').number_format = '#,##0'
    ws_vehicle.cell(row=10, column=6, value='=IF(B10=0,0,E10/B10)').number_format = '0.0%'
    
    # 列幅調整
    ws_vehicle.column_dimensions['A'].width = 15
    for col in range(2, 7):
        ws_vehicle.column_dimensions[chr(64 + col)].width = 15
    
    # =========================
    # 8. 資金繰り表
    # =========================
    ws_cashflow = wb.create_sheet("資金繰り表")
    ws_cashflow.sheet_view.showGridLines = False
    
    ws_cashflow['A1'] = "月次資金繰り表"
    ws_cashflow['A1'].font = title_font
    ws_cashflow.merge_cells('A1:N1')
    
    # ヘッダー
    ws_cashflow['A3'] = "項目"
    ws_cashflow['A3'].fill = header_fill
    ws_cashflow['A3'].font = header_font
    
    for col, month in enumerate(months[:12], 2):
        cell = ws_cashflow.cell(row=3, column=col, value=month)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # 資金繰り項目
    cf_items = ["月初残高", "入金合計", "出金合計", "月末残高"]
    for row_idx, item in enumerate(cf_items, 4):
        ws_cashflow.cell(row=row_idx, column=1, value=item).font = Font(bold=True)
    
    # 初期残高
    ws_cashflow['B4'] = 1000000
    ws_cashflow['B4'].number_format = '#,##0'
    
    # 数式（1月の例）
    ws_cashflow['B5'] = '=SUMIFS(資金繰り入力!E:E,資金繰り入力!B:B,1,資金繰り入力!C:C,"入金")'
    ws_cashflow['B6'] = '=SUMIFS(資金繰り入力!E:E,資金繰り入力!B:B,1,資金繰り入力!C:C,"出金")'
    ws_cashflow['B7'] = '=B4+B5-B6'
    
    # 条件付き書式（残高50万円未満で赤背景）
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws_cashflow.conditional_formatting.add('B7:M7',
        CellIsRule(operator='lessThan', formula=['500000'], fill=red_fill))
    
    # 列幅調整
    ws_cashflow.column_dimensions['A'].width = 15
    for col in range(2, 14):
        ws_cashflow.column_dimensions[chr(64 + col)].width = 12
    
    # =========================
    # 9. ダッシュボード
    # =========================
    ws_dashboard = wb.create_sheet("ダッシュボード", 0)  # 最初のシートに配置
    ws_dashboard.sheet_view.showGridLines = False
    
    ws_dashboard['A1'] = "経営ダッシュボード"
    ws_dashboard['A1'].font = Font(bold=True, size=16)
    ws_dashboard.merge_cells('A1:H1')
    
    # 現在の状況サマリー
    ws_dashboard['A3'] = "現在の経営状況"
    ws_dashboard['A3'].font = subtitle_font
    
    ws_dashboard['A5'] = "当月売上高:"
    ws_dashboard['B5'] = '=損益計算表!B5'
    ws_dashboard['B5'].number_format = '#,##0"円"'
    
    ws_dashboard['A6'] = "当月営業利益:"
    ws_dashboard['B6'] = '=損益計算表!B16'
    ws_dashboard['B6'].number_format = '#,##0"円"'
    
    ws_dashboard['A7'] = "当月営業利益率:"
    ws_dashboard['B7'] = '=損益計算表!B17'
    ws_dashboard['B7'].number_format = '0.0%'
    
    ws_dashboard['A8'] = "現在資金残高:"
    ws_dashboard['B8'] = '=資金繰り表!B7'
    ws_dashboard['B8'].number_format = '#,##0"円"'
    
    # 条件付き書式設定
    ws_dashboard.conditional_formatting.add('B6',
        CellIsRule(operator='lessThan', formula=['0'], font=red_font))
    ws_dashboard.conditional_formatting.add('B8',
        CellIsRule(operator='lessThan', formula=['500000'], fill=red_fill))
    
    # 警告メッセージエリア
    ws_dashboard['D3'] = "警告・注意事項"
    ws_dashboard['D3'].font = subtitle_font
    ws_dashboard['D3'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    ws_dashboard['D5'] = '=IF(損益計算表!B16<0,"⚠ 当月は赤字です","")'
    ws_dashboard['D5'].font = Font(color="FF0000", bold=True)
    
    ws_dashboard['D6'] = '=IF(資金繰り表!B7<500000,"⚠ 資金残高が50万円を下回っています","")'
    ws_dashboard['D6'].font = Font(color="FF0000", bold=True)
    
    # 車両別収支サマリー
    ws_dashboard['A11'] = "車両別収支TOP3"
    ws_dashboard['A11'].font = subtitle_font
    
    ws_dashboard['A13'] = "順位"
    ws_dashboard['B13'] = "車両番号"
    ws_dashboard['C13'] = "利益"
    for col in range(1, 4):
        ws_dashboard.cell(row=13, column=col).fill = header_fill
        ws_dashboard.cell(row=13, column=col).font = header_font
    
    # グラフプレースホルダー説明
    ws_dashboard['F3'] = "グラフ表示エリア"
    ws_dashboard['F3'].font = subtitle_font
    ws_dashboard['F5'] = "※ Excel上でグラフを挿入してください"
    ws_dashboard['F6'] = "1. 損益推移グラフ（月次）"
    ws_dashboard['F7'] = "2. 資金残高推移グラフ（月次）"
    ws_dashboard['F8'] = "3. 車両別収支グラフ（棒グラフ）"
    
    # 列幅調整
    ws_dashboard.column_dimensions['A'].width = 20
    ws_dashboard.column_dimensions['B'].width = 20
    ws_dashboard.column_dimensions['C'].width = 15
    ws_dashboard.column_dimensions['D'].width = 35
    
    # ワークブック保存
    wb.save("運送会社経営管理_v1.0.xlsx")
    print("✅ Excelファイルを作成しました: 運送会社経営管理_v1.0.xlsx")

# 実行
if __name__ == "__main__":
    create_transport_management_excel()
